# -*- coding: utf-8 -*-
"""
Основная программа для сравнения показателей по двум или трем периодам
Анализирует Excel файлы и рассчитывает приросты по клиентским менеджерам и клиентам
"""

import pandas as pd
import numpy as np
from typing import List, Dict, Tuple, Optional
from config import ANALYSIS_CONFIG, TEST_DATA_CONFIG, PROGRAM_MODES, IN_XLSX_DIR, TB_GOSB_CODES
from logger import logger
from test_data_generator import create_test_data


class PeriodComparison:
    """
    Класс для сравнения показателей по периодам
    Обрабатывает Excel файлы и рассчитывает приросты
    """
    
    def __init__(self):
        """
        Инициализация класса сравнения периодов
        Загружает конфигурацию и настраивает параметры
        """
        self.config = ANALYSIS_CONFIG
        # Определяем количество файлов на основе параметра use_file
        self.file_count = sum(1 for file_config in self.config['files'] if file_config.get('use_file', True))
        self.files_config = self.config['files']
        self.output_config = self.config['output']
        self.program_mode = PROGRAM_MODES['mode']
        self.aggregation_mode = self.config.get('aggregation_mode', 1)
        self.field_mapping = self.config.get('field_mapping', {})
        self.tb_gosb_codes = TB_GOSB_CODES
        
        # Словари для хранения данных из каждого файла
        self.data_frames = {}
        self.clients_data = {}
        self.managers_data = {}
        
        logger.debug(f"Инициализация с количеством файлов: {self.file_count}")
    
    def _get_files_for_mode(self):
        """
        Определяет какие файлы использовать в зависимости от режима работы
        
        Returns:
            list: Список файлов для обработки
        """
        if self.program_mode in [2, 4]:  # Режимы работы с тестовыми данными
            # Используем тестовые файлы, генерируем имена на основе обычных файлов
            from config import ANALYSIS_CONFIG
            used_files = [file_config for file_config in ANALYSIS_CONFIG['files'] if file_config.get('use_file', True)]
            files_to_use = []
            
            for i in range(self.file_count):
                if i < len(used_files):
                    # Генерируем имя тестового файла на основе обычного файла
                    import os
                    original_path = used_files[i]['path']
                    original_filename = os.path.basename(original_path)
                    test_filename = f"test_{original_filename}"
                    
                    file_config = {
                        'path': str(IN_XLSX_DIR / test_filename),
                        'sheet_name': 'Sheet1',
                        'columns': {
                            'Таб. номер': 'tab_number',
                            'КМ': 'fio',
                            'ТБ': 'tb',
                            'ГОСБ': 'gosb',
                            'ИНН': 'client_id',
                            'Клиент': 'client_name',
                            'ФОТ': 'value'
                        }
                    }
                else:
                    # Если файлов больше, чем в конфигурации, используем имя по умолчанию
                    file_config = {
                        'path': str(IN_XLSX_DIR / f'test_data_period{i+1}.xlsx'),
                        'sheet_name': 'Sheet1',
                        'columns': {
                            'Таб. номер': 'tab_number',
                            'КМ': 'fio',
                            'ТБ': 'tb',
                            'ГОСБ': 'gosb',
                            'ИНН': 'client_id',
                            'Клиент': 'client_name',
                            'ФОТ': 'value'
                        }
                    }
                files_to_use.append(file_config)
            
            return files_to_use
        else:  # Режимы работы с обычными данными
            # Используем основные файлы из конфигурации (только те, где use_file=True)
            return [file_config for file_config in self.files_config if file_config.get('use_file', True)]
    
    def _validate_tab_number(self, value) -> int:
        """
        Валидация и очистка табельного номера
        Формат: 8 знаков с лидирующими нулями
        Заменяет нечисловые значения на 70000000 (новые случаи) или 90000000 (серая зона)
        
        Args:
            value: Значение для валидации
            
        Returns:
            int: Валидный табельный номер в формате 8 знаков
        """
        try:
            # Преобразуем в строку и очищаем от лишних символов (включая апострофы)
            str_value = str(value).strip().replace("'", "")
            
            # Проверяем на специальные случаи
            if str_value.lower() in ['grey_zone', 'grey zone', 'greyzone']:
                logger.debug(f"Найдено значение 'grey_zone' в табельном номере, заменяем на 90000000")
                return 90000000
            
            if str_value in ['-', '', 'nan', 'None', 'null']:
                logger.debug(f"Найдено пустое или некорректное значение в табельном номере, заменяем на 70000000")
                return 70000000
            
            # Пытаемся преобразовать в число
            numeric_value = float(str_value)
            
            # Проверяем, что это целое число и положительное
            if numeric_value.is_integer() and numeric_value >= 0:
                # Преобразуем в 8-значный формат с лидирующими нулями
                tab_number = int(numeric_value)
                # Проверяем, что номер не превышает 8 знаков
                if tab_number > 99999999:
                    logger.debug(f"Табельный номер {value} превышает 8 знаков, заменяем на 70000000")
                    return 70000000
                return tab_number
            else:
                logger.debug(f"Табельный номер {value} не является положительным целым числом, заменяем на 70000000")
                return 70000000
                
        except (ValueError, TypeError):
            logger.debug(f"Не удалось преобразовать табельный номер '{value}' в число, заменяем на 70000000")
            return 70000000
    
    def _is_excluded_tab_number(self, tab_number: int) -> bool:
        """
        Проверяет, является ли табельный номер исключенным (8XXYYYYY или 9XXYYYYY)
        
        Args:
            tab_number: Табельный номер для проверки
            
        Returns:
            bool: True если табельный номер должен быть исключен
        """
        if tab_number == 0 or tab_number == 90000000:
            return False
        
        # Преобразуем в строку для проверки
        tab_str = str(tab_number).zfill(8)
        
        # Проверяем на паттерны 8XXYYYYY и 9XXYYYYY
        if len(tab_str) == 8:
            if tab_str[0] in ['8', '9']:
                return True
        
        return False
    
    def _validate_value(self, value) -> float:
        """
        Валидация и очистка показателя
        Заменяет нечисловые и пустые значения на 0
        
        Args:
            value: Значение для валидации
            
        Returns:
            float: Валидное числовое значение
        """
        try:
            # Проверяем на пустые значения
            if pd.isna(value) or value is None:
                logger.debug(f"Найдено пустое значение в показателе, заменяем на 0")
                return 0.0
            
            # Преобразуем в строку и очищаем
            str_value = str(value).strip()
            
            if str_value in ['', '-', 'nan', 'None', 'null']:
                logger.debug(f"Найдено некорректное значение в показателе, заменяем на 0")
                return 0.0
            
            # Пытаемся преобразовать в число
            numeric_value = float(str_value)
            
            # Проверяем на бесконечность и NaN
            if pd.isna(numeric_value) or not np.isfinite(numeric_value):
                logger.debug(f"Показатель {value} содержит NaN или бесконечность, заменяем на 0")
                return 0.0
            
            return numeric_value
            
        except (ValueError, TypeError):
            logger.debug(f"Не удалось преобразовать показатель '{value}' в число, заменяем на 0")
            return 0.0

    def load_excel_file(self, file_path: str, sheet_name: str, columns: Dict[str, str]) -> pd.DataFrame:
        """
        Загрузка данных из Excel файла
        
        Args:
            file_path (str): Путь к Excel файлу
            sheet_name (str): Название листа
            columns (Dict[str, str]): Словарь соответствия колонок
            
        Returns:
            pd.DataFrame: Загруженные данные
            
        Raises:
            FileNotFoundError: Если файл не найден
            Exception: При ошибке загрузки данных
        """
        try:
            logger.log_file_loading_start(file_path)
            logger.debug(f"Загружаем лист: {sheet_name}")
            
            # Загрузка данных из Excel файла
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            logger.debug(f"Исходный файл содержит {len(df)} строк и {len(df.columns)} колонок")
            
            # Переименование колонок согласно конфигурации
            df = df.rename(columns=columns)
            logger.log_file_columns_renamed(file_path, list(columns.keys()))
            logger.debug(f"Колонки переименованы: {columns}")
            
            # Очистка данных от пустых значений (после переименования)
            rows_before = len(df)
            required_columns = ['client_id', 'value']
            available_columns = [col for col in required_columns if col in df.columns]
            logger.debug(f"Доступные колонки для очистки: {available_columns}")
            if available_columns:
                df = df.dropna(subset=available_columns)
                logger.log_file_data_cleaned(file_path, rows_before, len(df))
            else:
                logger.debug("Нет колонок для очистки от пустых значений")
            
            # Валидация и очистка данных (после переименования)
            logger.debug("Начинаем валидацию и очистку данных")
            
            # Валидация табельных номеров
            if 'tab_number' in df.columns:
                logger.debug("Валидация табельных номеров")
                original_tab_count = len(df)
                original_invalid_tab = df['tab_number'].isna().sum() + (df['tab_number'].astype(str).str.strip().isin(['', '-', 'nan', 'None', 'null'])).sum()
                
                # Применяем валидацию к табельным номерам
                df['tab_number'] = df['tab_number'].apply(self._validate_tab_number)
                
                # Обработка табельных номеров с апострофами (после валидации)
                df['tab_number'] = df['tab_number'].astype(str).str.replace("'", "").str.zfill(8)
                
                logger.debug(f"Валидация табельных номеров завершена: {df['tab_number'].nunique()} уникальных")
                if original_invalid_tab > 0:
                    logger.info(f"Заменено {original_invalid_tab} некорректных табельных номеров на 70000000 и 90000000")
            
            # Валидация ID клиентов
            if 'client_id' in df.columns:
                logger.debug("Обработка ID клиентов")
                # Обработка ID клиентов с апострофами
                df['client_id'] = df['client_id'].astype(str).str.replace("'", "").str.zfill(20)
                logger.debug(f"Обработаны ID клиентов: {df['client_id'].nunique()} уникальных")
            
            # Валидация показателей
            if 'value' in df.columns:
                logger.debug("Валидация показателей")
                original_invalid_values = df['value'].isna().sum() + (df['value'].astype(str).str.strip().isin(['', '-', 'nan', 'None', 'null'])).sum()
                
                # Применяем валидацию к показателям
                df['value'] = df['value'].apply(self._validate_value)
                
                logger.debug(f"Валидация показателей завершена: среднее = {df['value'].mean():.2f}, сумма = {df['value'].sum():.2f}")
                if original_invalid_values > 0:
                    logger.info(f"Заменено {original_invalid_values} некорректных показателей на 0")
            
            logger.debug("Валидация и очистка данных завершена")
            
            logger.log_file_loaded(file_path)
            logger.log_file_data_processed(file_path, len(df))
            
            return df
            
        except FileNotFoundError:
            error_msg = f"Файл не найден: {file_path}"
            logger.log_file_load_error(file_path, error_msg)
            raise FileNotFoundError(error_msg)
        except Exception as e:
            error_msg = f"Ошибка загрузки файла {file_path}: {str(e)}"
            logger.log_error(error_msg)
            raise Exception(error_msg)
    
    def load_all_files(self) -> None:
        """
        Загрузка всех файлов согласно конфигурации
        Сохраняет данные в словарь data_frames
        """
        logger.debug("Начало загрузки всех файлов")
        
        # Получаем файлы в зависимости от режима работы
        files_to_load = self._get_files_for_mode()
        
        for i, file_config in enumerate(files_to_load):
            try:
                df = self.load_excel_file(
                    file_config['path'],
                    file_config['sheet_name'],
                    file_config['columns']
                )
                self.data_frames[f'period_{i+1}'] = df
                logger.debug(f"Файл периода {i+1} загружен успешно")
                
            except Exception as e:
                logger.log_error(f"Ошибка загрузки файла периода {i+1}: {str(e)}")
                raise
    
    def create_clients_base(self) -> pd.DataFrame:
        """
        Создание базы клиентов из уникальных идентификаторов
        Объединяет данные из всех периодов
        
        Returns:
            pd.DataFrame: База клиентов с данными по всем периодам
        """
        logger.debug("Создание базы клиентов")
        
        # Сбор всех уникальных ключей агрегации клиентов
        all_client_keys = set()
        for period_key, df in self.data_frames.items():
            for _, row in df.iterrows():
                client_key = self._get_client_aggregation_key(row)
                all_client_keys.add(client_key)
        
        # Создание базового DataFrame с клиентами
        clients_base = pd.DataFrame({'client_key': list(all_client_keys)})
        
        # Добавление данных по каждому периоду
        for i, (period_key, df) in enumerate(self.data_frames.items()):
            period_num = i + 1
            
            # Создание словаря для маппинга клиент -> данные с учетом агрегации
            period_data_list = []
            
            for _, row in df.iterrows():
                client_key = self._get_client_aggregation_key(row)
                period_data_list.append({
                    'client_key': client_key,
                    'client_id': row['client_id'],
                    'tab_number': row['tab_number'],
                    'fio': row['fio'],
                    'tb': row['tb'],
                    'gosb': row['gosb'],
                    'client_name': row['client_name'],
                    'value': row['value']
                })
            
            # Группировка по ключу агрегации и суммирование
            period_df = pd.DataFrame(period_data_list)
            period_data = period_df.groupby('client_key').agg({
                'client_id': 'first',
                'tab_number': 'first',
                'fio': 'first',
                'tb': 'first',
                'gosb': 'first',
                'client_name': 'first',
                'value': 'sum'  # Суммируем все значения по ключу агрегации
            }).reset_index()
            
            # Добавление колонок с данными периода
            clients_base = clients_base.merge(
                period_data[['client_key', 'client_id', 'tab_number', 'fio', 'tb', 'gosb', 'client_name', 'value']],
                on='client_key',
                how='left',
                suffixes=('', f'_period_{period_num}')
            )
            
            # Переименование колонок для ясности
            clients_base = clients_base.rename(columns={
                'tab_number': f'tab_number_period_{period_num}',
                'fio': f'fio_period_{period_num}',
                'tb': f'tb_period_{period_num}',
                'gosb': f'gosb_period_{period_num}',
                'client_name': f'client_name_period_{period_num}',
                'value': f'value_period_{period_num}'
            })
        
        # Заполнение пропущенных значений
        for i in range(1, self.file_count + 1):
            clients_base[f'tab_number_period_{i}'] = clients_base[f'tab_number_period_{i}'].fillna(0).astype(int)
            clients_base[f'value_period_{i}'] = clients_base[f'value_period_{i}'].fillna(0)
            clients_base[f'fio_period_{i}'] = clients_base[f'fio_period_{i}'].fillna('')
            clients_base[f'tb_period_{i}'] = clients_base[f'tb_period_{i}'].fillna('')
            clients_base[f'gosb_period_{i}'] = clients_base[f'gosb_period_{i}'].fillna('')
            clients_base[f'client_name_period_{i}'] = clients_base[f'client_name_period_{i}'].fillna('')
            
            # Для табельного номера 70000000 устанавливаем "-" в ТБ и ГОСБ
            mask_70000000 = clients_base[f'tab_number_period_{i}'] == 70000000
            clients_base.loc[mask_70000000, f'tb_period_{i}'] = '-'
            clients_base.loc[mask_70000000, f'gosb_period_{i}'] = '-'
            
            # Для табельного номера 90000000 устанавливаем "-" в ТБ и ГОСБ
            mask_90000000 = clients_base[f'tab_number_period_{i}'] == 90000000
            clients_base.loc[mask_90000000, f'tb_period_{i}'] = '-'
            clients_base.loc[mask_90000000, f'gosb_period_{i}'] = '-'
        
        # Определение итогового табельного номера
        clients_base['final_tab_number'] = 0
        clients_base['final_fio'] = ''
        clients_base['final_tb'] = ''
        clients_base['final_gosb'] = ''
        
        # Логика выбора итогового табельного с учетом правил агрегации
        for i in range(self.file_count, 0, -1):
            mask = (clients_base[f'tab_number_period_{i}'] != 0) & (clients_base['final_tab_number'] == 0)
            
            # Применяем правила выбора менеджера
            for idx in clients_base[mask].index:
                row = clients_base.loc[idx]
                client_aggregation_mode = self.aggregation_mode
                
                # Выбираем менеджера с наибольшим показателем в соответствии с правилами агрегации
                best_manager = None
                best_value = 0
                
                # Ищем менеджера с наибольшим показателем
                for j in range(1, self.file_count + 1):
                    if (row[f'tab_number_period_{j}'] != 0 and 
                        row[f'value_period_{j}'] > best_value):
                        
                        # Проверяем соответствие правилам агрегации
                        if self._check_manager_client_match(row, j, client_aggregation_mode):
                            # Проверяем, что табельный номер не исключен (8XXYYYYY или 9XXYYYYY)
                            tab_number = row[f'tab_number_period_{j}']
                            if not self._is_excluded_tab_number(tab_number):
                                best_manager = j
                                best_value = row[f'value_period_{j}']
                
                if best_manager:
                    clients_base.loc[idx, 'final_tab_number'] = row[f'tab_number_period_{best_manager}']
                    clients_base.loc[idx, 'final_fio'] = row[f'fio_period_{best_manager}']
                    clients_base.loc[idx, 'final_tb'] = row[f'tb_period_{best_manager}']
                    clients_base.loc[idx, 'final_gosb'] = row[f'gosb_period_{best_manager}']
                else:
                    # Если не найден подходящий менеджер (все исключены), ищем любого менеджера
                    fallback_manager = None
                    fallback_value = 0
                    
                    for j in range(1, self.file_count + 1):
                        if (row[f'tab_number_period_{j}'] != 0 and 
                            row[f'value_period_{j}'] > fallback_value):
                            
                            # Проверяем соответствие правилам агрегации
                            if self._check_manager_client_match(row, j, client_aggregation_mode):
                                fallback_manager = j
                                fallback_value = row[f'value_period_{j}']
                    
                    if fallback_manager:
                        clients_base.loc[idx, 'final_tab_number'] = row[f'tab_number_period_{fallback_manager}']
                        clients_base.loc[idx, 'final_fio'] = row[f'fio_period_{fallback_manager}']
                        clients_base.loc[idx, 'final_tb'] = row[f'tb_period_{fallback_manager}']
                        clients_base.loc[idx, 'final_gosb'] = row[f'gosb_period_{fallback_manager}']
        
        # Обработка серой зоны и прочих данных
        self._process_special_zones(clients_base)
        
        # Для итогового табельного номера 70000000 устанавливаем "-" в ТБ и ГОСБ
        mask_final_70000000 = clients_base['final_tab_number'] == 70000000
        clients_base.loc[mask_final_70000000, 'final_tb'] = '-'
        clients_base.loc[mask_final_70000000, 'final_gosb'] = '-'
        
        # Для итогового табельного номера 90000000 устанавливаем "-" в ТБ и ГОСБ
        mask_final_90000000 = clients_base['final_tab_number'] == 90000000
        clients_base.loc[mask_final_90000000, 'final_tb'] = '-'
        clients_base.loc[mask_final_90000000, 'final_gosb'] = '-'
        
        logger.debug(f"База клиентов создана: {len(clients_base)} уникальных клиентов")
        return clients_base
    
    def calculate_growth(self, clients_base: pd.DataFrame) -> pd.DataFrame:
        """
        Расчет приростов согласно формуле
        T-0 (первый файл) - текущий период
        T-1 (второй файл) - прошлый период  
        T-2 (третий файл) - позапрошлый период
        
        Args:
            clients_base (pd.DataFrame): База клиентов с данными по периодам
            
        Returns:
            pd.DataFrame: База клиентов с рассчитанными приростами
        """
        logger.debug("Начинаем расчет приростов")
        
        # Расчет прироста в зависимости от количества периодов
        if self.file_count == 2:
            # Прирост = T-0 - T-1 (текущий - прошлый)
            clients_base['growth'] = (
                clients_base['value_period_1'] - clients_base['value_period_2']
            )
        elif self.file_count == 3:
            # Прирост = ((T-0) - (T-1)) - ((T-1) - (T-2))
            period_0_1 = clients_base['value_period_1'] - clients_base['value_period_2']
            period_1_2 = clients_base['value_period_2'] - clients_base['value_period_3']
            clients_base['growth'] = period_0_1 - period_1_2
        else:
            raise ValueError(f"Неподдерживаемое количество периодов: {self.file_count}")
        
        logger.debug("Расчет приростов завершен")
        logger.debug(f"Приросты рассчитаны для {len(clients_base)} клиентов")
        
        return clients_base
    
    def create_managers_summary(self, clients_base: pd.DataFrame) -> pd.DataFrame:
        """
        Создание сводки по менеджерам
        
        Args:
            clients_base (pd.DataFrame): База клиентов с приростами
            
        Returns:
            pd.DataFrame: Сводка по менеджерам
        """
        logger.debug("Создание сводки по менеджерам")
        
        # Группировка по ключу агрегации менеджеров
        manager_aggregation_mode = self.aggregation_mode
        
        # Создаем ключ агрегации для менеджеров
        clients_base['manager_key'] = clients_base.apply(
            lambda row: self._get_manager_aggregation_key_from_final(row), axis=1
        )
        
        agg_dict = {
            'final_fio': 'first',
            'final_tb': 'first',
            'final_gosb': 'first',
            'value_period_1': 'sum',
            'value_period_2': 'sum',
            'growth': 'sum'
        }
        
        # Добавляем value_period_3 только если есть 3 периода
        if self.file_count == 3:
            agg_dict['value_period_3'] = 'sum'
        
        managers_summary = clients_base.groupby('manager_key').agg(agg_dict).reset_index()
        
        # Переименование колонок для соответствия выходному формату
        rename_dict = {
            'final_fio': 'fio',
            'final_tb': 'tb',
            'final_gosb': 'gosb',
            'value_period_1': 'value_1',
            'value_period_2': 'value_2',
            'growth': 'total_growth'
        }
        
        # Добавляем value_period_3 только если есть 3 периода
        if self.file_count == 3:
            rename_dict['value_period_3'] = 'value_3'
        
        managers_summary = managers_summary.rename(columns=rename_dict)
        
        # Добавляем колонку tab_number из manager_key
        managers_summary['tab_number'] = managers_summary['manager_key'].apply(
            lambda x: int(x.split('_')[0]) if '_' in x else int(x)
        )
        
        # Удаление колонки value_3 если только 2 периода
        if self.file_count == 2:
            managers_summary = managers_summary.drop(columns=['value_3'], errors='ignore')
        
        logger.debug(f"Сводка по менеджерам создана: {len(managers_summary)} уникальных менеджеров")
        return managers_summary
    
    def create_managers_deal_date_summary(self, clients_base: pd.DataFrame) -> pd.DataFrame:
        """
        Создание сводки по менеджерам с расчетом по дате сделки
        Для каждого КМ суммируются показатели клиентов, которые были закреплены за ним в каждом периоде
        
        Args:
            clients_base (pd.DataFrame): База клиентов с приростами
            
        Returns:
            pd.DataFrame: Сводка по менеджерам по дате сделки
        """
        logger.debug("Создание сводки по менеджерам по дате сделки")
        
        # Создаем словарь для хранения данных по менеджерам
        managers_data = {}
        
        # Обрабатываем каждый период отдельно
        for period in range(1, self.file_count + 1):
            period_key = f'period_{period}'
            if period_key in self.data_frames:
                df = self.data_frames[period_key]
                
                # Группируем по табельному номеру и суммируем показатели
                period_summary = df.groupby('tab_number').agg({
                    'fio': 'first',
                    'tb': 'first', 
                    'gosb': 'first',
                    'value': 'sum'
                }).reset_index()
                
                # Сохраняем данные для каждого менеджера
                for _, row in period_summary.iterrows():
                    tab_num = row['tab_number']
                    if tab_num not in managers_data:
                        managers_data[tab_num] = {
                            'tab_number': tab_num,
                            'fio': row['fio'],
                            'tb': row['tb'],
                            'gosb': row['gosb'],
                            'value_1': 0,
                            'value_2': 0,
                            'value_3': 0
                        }
                    
                    # Записываем показатель для соответствующего периода
                    if period == 1:
                        managers_data[tab_num]['value_1'] = row['value']
                    elif period == 2:
                        managers_data[tab_num]['value_2'] = row['value']
                    elif period == 3 and self.file_count == 3:
                        managers_data[tab_num]['value_3'] = row['value']
        
        # Преобразуем словарь в DataFrame
        managers_list = list(managers_data.values())
        managers_summary = pd.DataFrame(managers_list)
        
        # Расчет прироста по дате сделки
        if self.file_count == 2:
            # Прирост = T-0 - T-1 (текущий - прошлый)
            managers_summary['total_growth'] = (
                managers_summary['value_1'] - managers_summary['value_2']
            )
        elif self.file_count == 3:
            # Прирост = ((T-0) - (T-1)) - ((T-1) - (T-2))
            period_0_1 = managers_summary['value_1'] - managers_summary['value_2']
            period_1_2 = managers_summary['value_2'] - managers_summary['value_3']
            managers_summary['total_growth'] = period_0_1 - period_1_2
        
        logger.debug(f"Сводка по менеджерам по дате сделки создана: {len(managers_summary)} уникальных менеджеров")
        return managers_summary
    
    def create_output_file(self, clients_base: pd.DataFrame, managers_summary: pd.DataFrame, managers_deal_date_summary: pd.DataFrame = None) -> None:
        """
        Создание выходного Excel файла с результатами
        
        Args:
            clients_base (pd.DataFrame): База клиентов с приростами
            managers_summary (pd.DataFrame): Сводка по менеджерам
            managers_deal_date_summary (pd.DataFrame, optional): Сводка по менеджерам по дате сделки
        """
        # Формирование имени файла с временной меткой
        base_name = self.output_config['file_name']
        if self.output_config.get('add_timestamp', False):
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_file = f"{base_name}_{timestamp}.xlsx"
        else:
            output_file = f"{base_name}.xlsx"
        
        logger.log_output_creation_start(output_file)
        logger.debug(f"Создаем выходной файл: {output_file}")
        
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # Подготовка данных для листа клиентов
                columns_to_select = [
                    'client_id',
                    'client_name_period_1',
                    'value_period_1',
                    'value_period_2',
                    'growth',
                    'final_tab_number',
                    'final_fio',
                    'final_gosb',
                    'final_tb'
                ]
                
                # Добавляем колонку value_period_3 только если есть 3 периода
                if self.file_count == 3:
                    columns_to_select.insert(4, 'value_period_3')
                
                clients_output = clients_base[columns_to_select].copy()
                
                # Удаление None колонок
                clients_output = clients_output.dropna(axis=1, how='all')
                
                # Переименование колонок для читаемости
                column_mapping = {
                    'client_id': 'ID client',
                    'client_name_period_1': 'Client Name',
                    'value_period_1': 'val (T-0)',
                    'value_period_2': 'val (T-1)',
                    'value_period_3': 'val (T-2)',
                    'growth': 'Gain',
                    'final_tab_number': 'TN (final)',
                    'final_fio': 'ФИО КМ (final)',
                    'final_gosb': 'ГОСБ',
                    'final_tb': 'ТБ'
                }
                
                clients_output = clients_output.rename(columns=column_mapping)
                
                # Запись листа клиентов
                clients_output.to_excel(
                    writer,
                    sheet_name=self.output_config['sheets']['clients'],
                    index=False
                )
                
                # Подготовка данных для листа менеджеров
                managers_output = managers_summary.copy()
                
                # Переименование колонок для читаемости
                managers_column_mapping = {
                    'tab_number': 'TN (unic)',
                    'fio': 'ФИО',
                    'tb': 'ТБ',
                    'gosb': 'ГОСБ',
                    'value_1': 'val (T-0)',
                    'value_2': 'val (T-1)',
                    'value_3': 'val (T-2)',
                    'total_growth': 'Gain (total)'
                }
                
                managers_output = managers_output.rename(columns=managers_column_mapping)
                
                # Удаление колонки val (T-2) если только 2 периода
                if self.file_count == 2:
                    managers_output = managers_output.drop(columns=['val (T-2)'], errors='ignore')
                
                # Запись листа менеджеров
                managers_output.to_excel(
                    writer,
                    sheet_name=self.output_config['sheets']['managers'],
                    index=False
                )
                
                # Создание листа менеджеров по дате сделки (если есть данные)
                if managers_deal_date_summary is not None:
                    managers_deal_date_output = managers_deal_date_summary.copy()
                    
                    # Переименование колонок для читаемости
                    managers_deal_date_column_mapping = {
                        'tab_number': 'TN (unic)',
                        'fio': 'ФИО',
                        'tb': 'ТБ',
                        'gosb': 'ГОСБ',
                        'value_1': 'val (T-0)',
                        'value_2': 'val (T-1)',
                        'value_3': 'val (T-2)',
                        'total_growth': 'Gain (total)'
                    }
                    
                    managers_deal_date_output = managers_deal_date_output.rename(columns=managers_deal_date_column_mapping)
                    
                    # Удаление колонки val (T-2) если только 2 периода
                    if self.file_count == 2:
                        managers_deal_date_output = managers_deal_date_output.drop(columns=['val (T-2)'], errors='ignore')
                    
                    # Запись листа менеджеров по дате сделки
                    managers_deal_date_output.to_excel(
                        writer,
                        sheet_name=self.output_config['sheets']['managers_deal_date'],
                        index=False
                    )
                
                # Создание листа детализации клиентов
                clients_detail = self._create_clients_detail_sheet(clients_base)
                clients_detail.to_excel(
                    writer,
                    sheet_name='Детализация клиентов',
                    index=False
                )
            
            # Применение форматирования после записи всех данных
            self._apply_formatting_to_file(output_file)
            
            logger.log_output_created(output_file)
            logger.debug(f"Выходной файл создан: {output_file}")
            
        except Exception as e:
            error_msg = f"Ошибка создания выходного файла: {str(e)}"
            logger.log_error(error_msg)
            raise Exception(error_msg)
    
    def _create_clients_detail_sheet(self, clients_base: pd.DataFrame) -> pd.DataFrame:
        """
        Создание листа детализации клиентов с данными из всех файлов
        
        Args:
            clients_base (pd.DataFrame): База клиентов с приростами
            
        Returns:
            pd.DataFrame: Данные для листа детализации
        """
        logger.debug("Создание листа детализации клиентов")
        
        # Создаем список колонок для детализации
        detail_columns = [
            'client_id',
            'client_name_period_1',
            'final_tb',
            'final_gosb'
        ]
        
        # Добавляем данные по каждому периоду
        for i in range(1, self.file_count + 1):
            detail_columns.extend([
                f'tab_number_period_{i}',
                f'fio_period_{i}',
                f'tb_period_{i}',
                f'gosb_period_{i}',
                f'value_period_{i}'
            ])
        
        # Добавляем итоговые данные
        detail_columns.extend([
            'final_tab_number',
            'final_fio',
            'growth'
        ])
        
        # Создаем DataFrame с детализацией
        clients_detail = clients_base[detail_columns].copy()
        
        # Переименовываем колонки для читаемости
        column_mapping = {
            'client_id': 'ID клиента',
            'client_name_period_1': 'Наименование клиента',
            'final_tb': 'Итоговый ТБ',
            'final_gosb': 'Итоговый ГОСБ'
        }
        
        # Добавляем переименования для периодов
        for i in range(1, self.file_count + 1):
            period_name = f"T-{i-1}" if i > 1 else "T-0"
            column_mapping.update({
                f'tab_number_period_{i}': f'Таб. номер ({period_name})',
                f'fio_period_{i}': f'ФИО КМ ({period_name})',
                f'tb_period_{i}': f'ТБ ({period_name})',
                f'gosb_period_{i}': f'ГОСБ ({period_name})',
                f'value_period_{i}': f'Показатель ({period_name})'
            })
        
        # Добавляем переименования для итоговых данных
        column_mapping.update({
            'final_tab_number': 'Итоговый таб. номер',
            'final_fio': 'Итоговое ФИО КМ',
            'growth': 'Итоговый прирост'
        })
        
        # Применяем переименования
        clients_detail = clients_detail.rename(columns=column_mapping)
        
        # Сортируем по ID клиента для удобства
        clients_detail = clients_detail.sort_values('ID клиента')
        
        logger.debug(f"Создан лист детализации с {len(clients_detail)} записями")
        return clients_detail
    
    def _apply_formatting_to_file(self, file_path: str) -> None:
        """
        Применение форматирования к созданному Excel файлу
        
        Args:
            file_path: Путь к Excel файлу
        """
        logger.debug("Применение форматирования к файлу")
        
        try:
            from openpyxl import load_workbook
            from openpyxl.utils import get_column_letter
            
            # Загрузка файла
            wb = load_workbook(file_path)
            
            # Получение настроек форматирования
            formatting_config = self.output_config.get('formatting', {})
            
            # Форматирование листа клиентов
            if 'clients' in formatting_config:
                clients_sheet = wb[self.output_config['sheets']['clients']]
                self._format_sheet_columns(clients_sheet, formatting_config['clients'])
                self._apply_autofilter_and_freeze(clients_sheet)
            
            # Форматирование листа менеджеров
            if 'managers' in formatting_config:
                managers_sheet = wb[self.output_config['sheets']['managers']]
                self._format_sheet_columns(managers_sheet, formatting_config['managers'])
                self._apply_autofilter_and_freeze(managers_sheet)
            
            # Форматирование листа менеджеров по дате сделки
            if 'managers_deal_date' in formatting_config:
                managers_deal_date_sheet = wb[self.output_config['sheets']['managers_deal_date']]
                self._format_sheet_columns(managers_deal_date_sheet, formatting_config['managers_deal_date'])
                self._apply_autofilter_and_freeze(managers_deal_date_sheet)
            
            # Форматирование листа детализации клиентов
            if 'Детализация клиентов' in wb.sheetnames:
                clients_detail_sheet = wb['Детализация клиентов']
                # Применяем форматирование колонок для листа детализации
                if 'clients_detail' in formatting_config:
                    self._format_sheet_columns(clients_detail_sheet, formatting_config['clients_detail'])
                self._apply_autofilter_and_freeze(clients_detail_sheet)
                # Автоподбор ширины колонок
                for column in clients_detail_sheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)  # Максимум 50 символов
                    clients_detail_sheet.column_dimensions[column_letter].width = adjusted_width
            
            # Сохранение файла с форматированием
            wb.save(file_path)
            wb.close()
            
            logger.debug("Форматирование применено успешно")
            
        except Exception as e:
            logger.log_error(f"Ошибка применения форматирования: {str(e)}")
            # Не прерываем выполнение, так как форматирование не критично
    
    def _format_sheet_columns(self, sheet, column_formats: dict) -> None:
        """
        Форматирование колонок конкретного листа
        
        Args:
            sheet: Лист Excel файла
            column_formats: Словарь с настройками форматирования колонок
        """
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter
        
        # Создание стилей для разных типов данных
        number_font = Font(name='Arial', size=10)
        number_alignment = Alignment(horizontal='right')
        
        text_font = Font(name='Arial', size=10)
        text_alignment = Alignment(horizontal='left')
        
        # Получение заголовков для поиска колонок
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # Применение форматирования к каждой колонке
        for col_name, format_config in column_formats.items():
            # Поиск индекса колонки по имени
            col_idx = None
            for i, header in enumerate(headers, 1):
                if header == col_name:
                    col_idx = i
                    break
            
            if col_idx is None:
                logger.debug(f"Колонка '{col_name}' не найдена в листе")
                continue
                
            col_letter = get_column_letter(col_idx)
            logger.debug(f"Форматирование колонки {col_name} ({col_letter})")
            
            # Применение стиля и формата
            if format_config['type'] == 'number':
                # Применение числового формата
                for row in range(2, sheet.max_row + 1):
                    cell = sheet[f"{col_letter}{row}"]
                    cell.number_format = format_config.get('format', '#,##0.00')
                    cell.font = number_font
                    cell.alignment = number_alignment
            elif format_config['type'] == 'text':
                # Применение текстового стиля
                for row in range(2, sheet.max_row + 1):
                    cell = sheet[f"{col_letter}{row}"]
                    cell.font = text_font
                    cell.alignment = text_alignment
            elif format_config['type'] == 'text_padded':
                # Применение текстового стиля с дополнением нулями
                pad_length = int(format_config.get('format', '8'))
                pad_char = format_config.get('pad_char', '0')
                
                for row in range(2, sheet.max_row + 1):
                    cell = sheet[f"{col_letter}{row}"]
                    # Дополняем значение нулями до нужной длины
                    if cell.value is not None:
                        cell.value = str(cell.value).zfill(pad_length)
                    cell.font = text_font
                    cell.alignment = text_alignment
        
        # Автоподбор ширины колонок
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
            sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_autofilter_and_freeze(self, sheet) -> None:
        """
        Применение автофильтра и закрепления первой строки
        
        Args:
            sheet: Лист Excel файла
        """
        try:
            from openpyxl.utils import get_column_letter
            
            # Применение автофильтра ко всем данным
            if sheet.max_row > 1:
                sheet.auto_filter.ref = f"A1:{get_column_letter(sheet.max_column)}{sheet.max_row}"
            
            # Закрепление первой строки
            sheet.freeze_panes = "A2"
            
            logger.debug(f"Автофильтр и закрепление применены к листу {sheet.title}")
            
        except Exception as e:
            logger.log_error(f"Ошибка применения автофильтра и закрепления: {str(e)}")
    
    def _get_client_aggregation_key(self, row) -> str:
        """
        Получение ключа агрегации для клиента в зависимости от режима
        
        Args:
            row: Строка данных
            
        Returns:
            str: Ключ агрегации
        """
        client_aggregation_mode = self.aggregation_mode
        
        if client_aggregation_mode == 1:
            # Агрегация только по client_id
            return str(row['client_id'])
        elif client_aggregation_mode == 2:
            # Агрегация по client_id + tb
            return f"{row['client_id']}_{row['tb']}"
        elif client_aggregation_mode == 3:
            # Агрегация по client_id + tb + gosb
            return f"{row['client_id']}_{row['tb']}_{row['gosb']}"
        else:
            # По умолчанию по client_id
            return str(row['client_id'])
    
    def _get_manager_aggregation_key(self, row) -> str:
        """
        Получение ключа агрегации для менеджера в зависимости от режима
        
        Args:
            row: Строка данных
            
        Returns:
            str: Ключ агрегации
        """
        manager_aggregation_mode = self.aggregation_mode
        
        if manager_aggregation_mode == 1:
            # Агрегация только по tab_number
            return str(row['tab_number'])
        elif manager_aggregation_mode == 2:
            # Агрегация по tab_number + tb
            return f"{row['tab_number']}_{row['tb']}"
        elif manager_aggregation_mode == 3:
            # Агрегация по tab_number + tb + gosb
            return f"{row['tab_number']}_{row['tb']}_{row['gosb']}"
        else:
            # По умолчанию по tab_number
            return str(row['tab_number'])
    
    def _get_tb_code_from_name(self, tb_name: str) -> int:
        """
        Получение кода ТБ по названию с приоритетом поиска
        
        Args:
            tb_name: Название ТБ
            
        Returns:
            int: Код ТБ или 0 если не найден
        """
        if not tb_name or tb_name == '-':
            return 0
            
        # Приоритет поиска: код -> короткое имя -> полное имя
        for tb_code, tb_info in self.tb_gosb_codes['tb_codes'].items():
            # 1. Сначала ищем по коду (если переданное значение - число)
            try:
                if int(tb_name) == tb_code:
                    return tb_code
            except (ValueError, TypeError):
                pass
            
            # 2. Затем по короткому имени
            if tb_name == tb_info['short_name']:
                return tb_code
                
            # 3. Затем по полному имени
            if tb_name == tb_info['full_name']:
                return tb_code
                
            # 4. Частичное совпадение с полным именем
            if tb_info['full_name'] and tb_name in tb_info['full_name']:
                return tb_code
                
        return 0
    
    def _get_gosb_code_from_name(self, gosb_name: str, tb_code: int) -> int:
        """
        Получение кода ГОСБ по названию и коду ТБ с приоритетом поиска
        
        Args:
            gosb_name: Название ГОСБ
            tb_code: Код ТБ
            
        Returns:
            int: Код ГОСБ или 0 если не найден
        """
        if not gosb_name or gosb_name == '-':
            return 0
            
        # Приоритет поиска: код -> точное совпадение -> частичное совпадение
        for (tb, gosb), name in self.tb_gosb_codes['gosb_codes'].items():
            if tb == tb_code:
                # 1. Сначала ищем по коду (если переданное значение - число)
                try:
                    if int(gosb_name) == gosb:
                        return gosb
                except (ValueError, TypeError):
                    pass
                
                # 2. Затем точное совпадение с названием
                if gosb_name == name:
                    return gosb
                    
                # 3. Затем частичное совпадение
                if gosb_name in name:
                    return gosb
                    
        return 0
    
    def _generate_grey_zone_tab_number(self, tb_code: int = 0, gosb_code: int = 0) -> int:
        """
        Генерация табельного номера для серой зоны
        
        Args:
            tb_code: Код ТБ
            gosb_code: Код ГОСБ
            
        Returns:
            int: Табельный номер для серой зоны
        """
        if tb_code == 0:
            return 90000000  # Базовая серая зона
        elif gosb_code == 0:
            return 90000000 + tb_code * 1000  # 9XX00000
        else:
            # 9XXYYYYY где XX - код ТБ, YYYYY - код ГОСБ (до 6 знаков)
            gosb_padded = str(gosb_code).zfill(6)
            return int(f"9{tb_code:02d}{gosb_padded}")
    
    def _generate_other_tab_number(self, tb_code: int = 0, gosb_code: int = 0) -> int:
        """
        Генерация табельного номера для прочих данных
        
        Args:
            tb_code: Код ТБ
            gosb_code: Код ГОСБ
            
        Returns:
            int: Табельный номер для прочих данных
        """
        if tb_code == 0:
            return 80000000  # Базовая прочая зона
        elif gosb_code == 0:
            return 80000000 + tb_code * 1000  # 8XX00000
        else:
            # 8XXYYYYY где XX - код ТБ, YYYYY - код ГОСБ (до 6 знаков)
            gosb_padded = str(gosb_code).zfill(6)
            return int(f"8{tb_code:02d}{gosb_padded}")
    
    def _check_manager_client_match(self, row, period_num: int, client_aggregation_mode: int) -> bool:
        """
        Проверка соответствия менеджера и клиента по правилам агрегации
        
        Args:
            row: Строка данных клиента
            period_num: Номер периода
            client_aggregation_mode: Режим агрегации клиентов
            
        Returns:
            bool: True если менеджер соответствует клиенту по правилам
        """
        if client_aggregation_mode == 1:
            # Агрегация только по client_id - все менеджеры подходят
            return True
        elif client_aggregation_mode == 2:
            # Агрегация по client_id + tb - проверяем ТБ
            return row[f'tb_period_{period_num}'] == row['final_tb']
        elif client_aggregation_mode == 3:
            # Агрегация по client_id + tb + gosb - проверяем ТБ и ГОСБ
            return (row[f'tb_period_{period_num}'] == row['final_tb'] and 
                    row[f'gosb_period_{period_num}'] == row['final_gosb'])
        else:
            return True
    
    def _get_manager_aggregation_key_from_final(self, row) -> str:
        """
        Получение ключа агрегации для менеджера из итоговых данных
        
        Args:
            row: Строка данных
            
        Returns:
            str: Ключ агрегации
        """
        manager_aggregation_mode = self.aggregation_mode
        
        if manager_aggregation_mode == 1:
            # Агрегация только по final_tab_number
            return str(row['final_tab_number'])
        elif manager_aggregation_mode == 2:
            # Агрегация по final_tab_number + final_tb
            return f"{row['final_tab_number']}_{row['final_tb']}"
        elif manager_aggregation_mode == 3:
            # Агрегация по final_tab_number + final_tb + final_gosb
            return f"{row['final_tab_number']}_{row['final_tb']}_{row['final_gosb']}"
        else:
            # По умолчанию по final_tab_number
            return str(row['final_tab_number'])
    
    def _process_special_zones(self, clients_base: pd.DataFrame) -> None:
        """
        Обработка серой зоны и прочих данных
        
        Args:
            clients_base: База клиентов для обработки
        """
        logger.debug("Обработка серой зоны и прочих данных")
        
        # Создаем словари для группировки данных по ТБ и ГОСБ
        tb_groups = {}
        gosb_groups = {}
        
        # Группируем данные по ТБ и ГОСБ
        for _, row in clients_base.iterrows():
            tb_name = row['final_tb']
            gosb_name = row['final_gosb']
            
            if tb_name and tb_name != '-':
                tb_code = self._get_tb_code_from_name(tb_name)
                if tb_code > 0:
                    if tb_code not in tb_groups:
                        tb_groups[tb_code] = []
                    tb_groups[tb_code].append(row)
                    
                    if gosb_name and gosb_name != '-':
                        gosb_code = self._get_gosb_code_from_name(gosb_name, tb_code)
                        if gosb_code > 0:
                            gosb_key = (tb_code, gosb_code)
                            if gosb_key not in gosb_groups:
                                gosb_groups[gosb_key] = []
                            gosb_groups[gosb_key].append(row)
        
        # Обрабатываем серую зону (90000000)
        self._process_grey_zone(clients_base, tb_groups, gosb_groups)
        
        # Обрабатываем прочие данные (80000000)
        self._process_other_zone(clients_base, tb_groups, gosb_groups)
    
    def _process_grey_zone(self, clients_base: pd.DataFrame, tb_groups: dict, gosb_groups: dict) -> None:
        """
        Обработка серой зоны (90000000)
        
        Args:
            clients_base: База клиентов
            tb_groups: Группы по ТБ
            gosb_groups: Группы по ГОСБ
        """
        # Базовая серая зона (90000000)
        grey_zone_data = self._create_special_zone_entry(
            tab_number=90000000,
            fio="Серая зона",
            tb="-",
            gosb="-",
            clients_base=clients_base
        )
        
        if grey_zone_data is not None:
            clients_base.loc[len(clients_base)] = grey_zone_data
        
        # Серая зона по ТБ (9XX00000)
        for tb_code, rows in tb_groups.items():
            tab_number = self._generate_grey_zone_tab_number(tb_code, 0)
            tb_name = self.tb_gosb_codes['tb_codes'][tb_code]['short_name']
            
            grey_zone_tb_data = self._create_special_zone_entry(
                tab_number=tab_number,
                fio=f"Серая зона {tb_name}",
                tb=tb_name,
                gosb="-",
                clients_base=clients_base,
                filter_rows=rows
            )
            
            if grey_zone_tb_data is not None:
                clients_base.loc[len(clients_base)] = grey_zone_tb_data
        
        # Серая зона по ТБ+ГОСБ (9XXYYYYY)
        for (tb_code, gosb_code), rows in gosb_groups.items():
            tab_number = self._generate_grey_zone_tab_number(tb_code, gosb_code)
            tb_name = self.tb_gosb_codes['tb_codes'][tb_code]['short_name']
            gosb_name = self.tb_gosb_codes['gosb_codes'][(tb_code, gosb_code)]
            
            grey_zone_gosb_data = self._create_special_zone_entry(
                tab_number=tab_number,
                fio=f"Серая зона {tb_name} {gosb_name}",
                tb=tb_name,
                gosb=gosb_name,
                clients_base=clients_base,
                filter_rows=rows
            )
            
            if grey_zone_gosb_data is not None:
                clients_base.loc[len(clients_base)] = grey_zone_gosb_data
    
    def _process_other_zone(self, clients_base: pd.DataFrame, tb_groups: dict, gosb_groups: dict) -> None:
        """
        Обработка прочих данных (80000000)
        
        Args:
            clients_base: База клиентов
            tb_groups: Группы по ТБ
            gosb_groups: Группы по ГОСБ
        """
        # Базовая прочая зона (80000000)
        other_zone_data = self._create_special_zone_entry(
            tab_number=80000000,
            fio="Прочее",
            tb="-",
            gosb="-",
            clients_base=clients_base
        )
        
        if other_zone_data is not None:
            clients_base.loc[len(clients_base)] = other_zone_data
        
        # Прочая зона по ТБ (8XX00000)
        for tb_code, rows in tb_groups.items():
            tab_number = self._generate_other_tab_number(tb_code, 0)
            tb_name = self.tb_gosb_codes['tb_codes'][tb_code]['short_name']
            
            other_zone_tb_data = self._create_special_zone_entry(
                tab_number=tab_number,
                fio=f"Прочее {tb_name}",
                tb=tb_name,
                gosb="-",
                clients_base=clients_base,
                filter_rows=rows
            )
            
            if other_zone_tb_data is not None:
                clients_base.loc[len(clients_base)] = other_zone_tb_data
        
        # Прочая зона по ТБ+ГОСБ (8XXYYYYY)
        for (tb_code, gosb_code), rows in gosb_groups.items():
            tab_number = self._generate_other_tab_number(tb_code, gosb_code)
            tb_name = self.tb_gosb_codes['tb_codes'][tb_code]['short_name']
            gosb_name = self.tb_gosb_codes['gosb_codes'][(tb_code, gosb_code)]
            
            other_zone_gosb_data = self._create_special_zone_entry(
                tab_number=tab_number,
                fio=f"Прочее {tb_name} {gosb_name}",
                tb=tb_name,
                gosb=gosb_name,
                clients_base=clients_base,
                filter_rows=rows
            )
            
            if other_zone_gosb_data is not None:
                clients_base.loc[len(clients_base)] = other_zone_gosb_data
    
    def _create_special_zone_entry(self, tab_number: int, fio: str, tb: str, gosb: str, 
                                 clients_base: pd.DataFrame, filter_rows: list = None) -> dict:
        """
        Создание записи для специальной зоны
        
        Args:
            tab_number: Табельный номер
            fio: ФИО
            tb: ТБ
            gosb: ГОСБ
            clients_base: База клиентов
            filter_rows: Отфильтрованные строки (опционально)
            
        Returns:
            dict: Данные для записи или None если нет данных
        """
        if filter_rows is None:
            # Берем все строки
            filtered_data = clients_base
        else:
            # Фильтруем по переданным строкам
            filtered_data = clients_base[clients_base.index.isin([r.name for r in filter_rows])]
        
        if len(filtered_data) == 0:
            return None
        
        # Суммируем данные
        entry = {
            'client_key': f"special_{tab_number}",
            'final_tab_number': tab_number,
            'final_fio': fio,
            'final_tb': tb,
            'final_gosb': gosb,
            'final_client_name': f"Специальная зона {tab_number}"
        }
        
        # Суммируем показатели по периодам
        for i in range(1, self.file_count + 1):
            value_col = f'value_period_{i}'
            if value_col in filtered_data.columns:
                entry[value_col] = filtered_data[value_col].sum()
            else:
                entry[value_col] = 0
        
        # Суммируем прирост
        if 'growth' in filtered_data.columns:
            entry['growth'] = filtered_data['growth'].sum()
        else:
            entry['growth'] = 0
        
        return entry
    
    def run_analysis(self) -> None:
        """
        Основной метод для запуска анализа
        Выполняет полный цикл обработки данных в зависимости от режима
        """
        try:
            logger.log_analysis_start()
            
            # Режим 1: Просто сгенерировать тест данные
            if self.program_mode == 1:
                logger.info("Режим 1: Генерация тестовых данных")
                self._generate_test_data_only()
                return
            
            # Режим 2: Посчитать на тест данных
            elif self.program_mode == 2:
                logger.info("Режим 2: Анализ тестовых данных")
                self._run_analysis_on_test_data()
                return
            
            # Режим 3: Посчитать на обычных данных
            elif self.program_mode == 3:
                logger.info("Режим 3: Анализ обычных данных")
                self._run_analysis_on_normal_data()
                return
            
            # Режим 4: Сгенерировать и посчитать на тест данных сразу
            elif self.program_mode == 4:
                logger.info("Режим 4: Генерация и анализ тестовых данных")
                self._generate_and_analyze_test_data()
                return
            
            else:
                raise ValueError(f"Неподдерживаемый режим работы: {self.program_mode}")
                
        except Exception as e:
            logger.log_error(f"Ошибка в процессе анализа: {str(e)}")
            raise
    
    def _generate_test_data_only(self) -> None:
        """Режим 1: Просто сгенерировать тест данные"""
        from test_data_generator import TestDataGenerator
        generator = TestDataGenerator()
        generator.create_test_files()
        logger.info("Тестовые данные сгенерированы успешно")
    
    def _run_analysis_on_test_data(self) -> None:
        """Режим 2: Посчитать на тест данных"""
        # Загрузка всех файлов
        self.load_all_files()
        
        # Создание базы клиентов
        clients_base = self.create_clients_base()
        logger.log_clients_base_created(len(clients_base))
        
        # Расчет приростов
        clients_base = self.calculate_growth(clients_base)
        logger.log_growth_calculated(len(clients_base))
        
        # Создание сводки по менеджерам
        managers_summary = self.create_managers_summary(clients_base)
        logger.log_managers_summary_created(len(managers_summary))
        
        # Создание сводки по менеджерам по дате сделки
        managers_deal_date_summary = self.create_managers_deal_date_summary(clients_base)
        logger.log_managers_deal_date_created(len(managers_deal_date_summary))
        
        # Создание выходного файла
        self.create_output_file(clients_base, managers_summary, managers_deal_date_summary)
        logger.log_analysis_complete()
    
    def _run_analysis_on_normal_data(self) -> None:
        """Режим 3: Посчитать на обычных данных"""
        # Загрузка всех файлов
        self.load_all_files()
        
        # Создание базы клиентов
        clients_base = self.create_clients_base()
        logger.log_clients_base_created(len(clients_base))
        
        # Расчет приростов
        clients_base = self.calculate_growth(clients_base)
        logger.log_growth_calculated(len(clients_base))
        
        # Создание сводки по менеджерам
        managers_summary = self.create_managers_summary(clients_base)
        logger.log_managers_summary_created(len(managers_summary))
        
        # Создание сводки по менеджерам по дате сделки
        managers_deal_date_summary = self.create_managers_deal_date_summary(clients_base)
        logger.log_managers_deal_date_created(len(managers_deal_date_summary))
        
        # Создание выходного файла
        self.create_output_file(clients_base, managers_summary, managers_deal_date_summary)
        logger.log_analysis_complete()
    
    def _generate_and_analyze_test_data(self) -> None:
        """Режим 4: Сгенерировать и посчитать на тест данных сразу"""
        # Сначала генерируем тестовые данные
        self._generate_test_data_only()
        
        # Затем анализируем их
        self._run_analysis_on_test_data()


def check_and_create_test_data():
    """
    Проверяет настройку создания тестовых данных и создает их при необходимости
    """
    logger.info("Создание тестовых данных...")
    
    # Удаление старых тестовых файлов
    import os
    from config import ANALYSIS_CONFIG
    
    # Получаем список файлов, которые используются (use_file=True)
    used_files = [file_config for file_config in ANALYSIS_CONFIG['files'] if file_config.get('use_file', True)]
    
    # Создаем список тестовых файлов на основе обычных файлов с префиксом "test_"
    test_files = []
    for file_config in used_files:
        import os
        original_path = file_config['path']
        original_filename = os.path.basename(original_path)
        test_filename = f"test_{original_filename}"
        test_files.append(IN_XLSX_DIR / test_filename)
    
    deleted_files = []
    for file_path in test_files:
        if file_path.exists():
            file_path.unlink()
            deleted_files.append(str(file_path))
            logger.debug(f"Удален старый тестовый файл: {file_path}")
    
    if deleted_files:
        logger.log_test_files_deleted(deleted_files)
    
    # Создание новых тестовых данных
    logger.info("Создание новых тестовых данных...")
    success = create_test_data()
    
    if success:
        created_files = [str(f) for f in test_files]
        logger.log_test_files_created(created_files)
        logger.info("Тестовые данные созданы успешно")
    else:
        logger.error("Ошибка создания тестовых данных")
        return False
    
    return True


def main():
    """
    Главная функция программы
    Запускает анализ периодов
    """
    try:
        # Создание экземпляра класса анализа
        analyzer = PeriodComparison()
        
        # Проверка и создание тестовых данных при необходимости (только для режимов 1 и 4)
        if analyzer.program_mode in [1, 4]:
            if not check_and_create_test_data():
                return 1
        
        # Запуск анализа
        analyzer.run_analysis()
        
        logger.log_program_end()
        # Получаем актуальное имя файла из конфигурации
        base_name = analyzer.output_config['file_name']
        if analyzer.output_config.get('add_timestamp', False):
            from datetime import datetime
            timestamp = datetime.now().strftime("%Y%m%d_%H%M")
            output_file = f"{base_name}_{timestamp}.xlsx"
        else:
            output_file = f"{base_name}.xlsx"
        logger.info(f"Анализ завершен успешно. Результаты сохранены в файл {output_file}")
        
    except Exception as e:
        logger.log_critical_error(str(e))
        print(f"Произошла ошибка: {str(e)}")
        return 1
    
    return 0


if __name__ == "__main__":
    exit(main())
